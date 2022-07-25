# _*_ coding:utf-8 _*_
# !/usr/bin/python3.7
# @Email:plz0927@163.com
# @Author: Russ


import os
import ast
import openpyxl
import yaml

"""
1、实现读取Excel的用例数据， 供excel_to_yaml模块使用
2、实现从yaml文件里读取用例数据，供PyTest使用
"""


# 从Excel读取用例数据


class ReadExcel:
    case_file_path = "../data/excel/"

    def __init__(self):
        self.sh = None

    def read_excel_2(self):
        """遍历固定excel固定路径下每一个excel文件，然后遍历每一个sheet, 然后读取excel用例，读取所有测试用例"""
        case_file_list = os.listdir('../TestExcel/')
        # 用例编号初始化
        case_sequence = 0
        # 遍历每个表单
        terminal_cases = {}
        # 把每一条case的key保存下来
        case_keys = []
        for case_file in case_file_list:
            file = self.case_file_path + case_file
            wb = openpyxl.load_workbook(file)

            # 得到所有表单名
            sheets = wb.sheetnames

            for i in sheets:
                # 定义操作的excel表单对象
                sh = wb[i]
                # 获取该表单的行数
                rows = list(sh.rows)
                titles = []
                # 把第一行的字段添加到列表
                for t in rows[0]:
                    title = t.value
                    titles.append(title)

                for row in rows[1:]:
                    # case = []
                    if row[-1].value == 1:  # 用来判断用例后续操作是否执行
                        # 遍历每一行的每个字段
                        num = 0
                        dict_cases = {}  # 用来存放一行用例数据
                        cases = []
                        # 一行的用例数据处理完成，得到列表，里面是字典（具体到每个字段）
                        for r in row:
                            temp_1 = r.value
                            # 判断字段是否属于字典类型
                            if titles[num] in ['body', 'excepted', 'headers', 'depend']:
                                # 如果该字段的值不是空，就把该字段值转为字典
                                if r.value is not None:
                                    temp_2 = ast.literal_eval(r.value)
                                    dict_cases[titles[num]] = temp_2
                                # 为空的话，就直接添加到字典dict_cases里
                                else:
                                    dict_cases[titles[num]] = temp_1
                            # 不是属于字典类型的，直接添加到字典dict_cases里
                            else:
                                dict_cases[titles[num]] = temp_1
                            num += 1
                        # 添加模块名，模块名：sheet命名
                        dict_cases['module_name'] = i
                        # 一条用例处理完成，得到列表
                        cases.append(dict_cases)

                        # 用例编号递增
                        case_sequence += 1
                        case_key = 'case' + str(case_sequence)

                        # 最终得到所有sheets的一个字典，key为case,value为列表，里面是字典，一个字典里是一条用例：
                        # {'case1': [{'case_id': 'case_001', 'interface':'任务查询',....}], 'case2':['']}
                        terminal_cases[case_key] = cases
                        case_keys.append(case_key)
                wb.close()
        return terminal_cases, case_keys

    def read_excel_3(self, file_name):
        excel_file = self.case_file_path + file_name
        wb = openpyxl.load_workbook(excel_file)
        # 得到所有的表单名
        sheets = wb.sheetnames
        # 遍历每个表单
        terminal_cases = {}

        # 用例编号初始化
        case_sequence = 0
        for i in sheets:
            # 定义操作excel表单对象
            self.sh = wb[i]
            # 获取该表单的行数
            rows = list(self.sh.rows)
            titles = []
            # 把第一行的字段添加到列表
            for t in rows[0]:
                title = t.value
                titles.append(title)

            for row in rows[1:]:
                if row[-1].value == 1:  # 【用来判断用例后续操作是否执行】
                    # 遍历每一行的每个字段
                    num = 0
                    dict_cases = {}  # 用来存放一行用例数据
                    cases = []
                    # 一行的用例数据处理完成，得到列表，里面是字典（具体到每个字段）
                    for r in row:
                        temp_1 = r.value
                        # 判断字段是否是属于字典类型的，
                        if titles[num] in ['body', 'excepted', 'headers', 'depend']:
                            # 如果该字段的值不是空，就把该字段值转为字典
                            if r.value is not None:
                                temp_2 = ast.literal_eval(r.value)
                                dict_cases[titles[num]] = temp_2
                            # 为空的话，就直接添加到字典dict_cases里
                            else:
                                dict_cases[titles[num]] = temp_1
                        # 不是属于字典类型的，直接添加到字典dict_cases里
                        else:
                            dict_cases[titles[num]] = temp_1
                        num += 1
                    # 添加模块名，模块名：sheet命名
                    dict_cases['module_name'] = i
                    # 一条用例处理完成，得到列表
                    cases.append(dict_cases)

                    # 用例编号递增
                    case_sequence += 1

                    # 最终得到所有sheets的一个字典，key为case,value为列表，里面是字典，一个字典里是一条用例：
                    # {'case1': [{'case_id': 'case_001', 'interface':'任务查询',....}], 'case2':['']}
                    terminal_cases['case' + str(case_sequence)] = cases
        wb.close()
        return terminal_cases


# 从数据库表读取用例数据
class ReadDatabase:
    def __init__(self, db_init, sql):
        self.db_init = db_init
        self.sql = sql

    def get_data(self):
        # 执行sql
        self.db_init.execute(self.sql)
        # 返回的就是一个列表，里面每个元素是个元组
        result = self.db_init.fetchall()
        return result


# 从yaml文件读取用例数据,供PyTest参数化使用
class ReadYaml:
    def __init__(self, filename):
        with open(filename, 'r', encoding='utf-8') as f:
            self.data = yaml.load(f, Loader=yaml.FullLoader)

    def get_yaml_data(self):
        cases = self.data
        return cases


# if __name__ == '__main__':
#     # case = ReadExcel().read_excel_2()[0]
#     yaml_file = '../output/test_case.yml'
#     cont = ReadYaml(yaml_file).get_yaml_data()
#     print(cont)
# -------------------------------------------------------------
#     cases = ReadExcel().read_excel_2()[0]
#     count = ReadExcel().read_excel_2()[-1]
#     print(cases)
#     print(count)
